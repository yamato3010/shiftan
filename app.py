import os
import glob
import datetime
from flask.scaffold import _matching_loader_thinks_module_is_package
import jpholiday
import numpy as np
from ortoolpy import addbinvars
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from flask import Flask, render_template
from flask import request, send_file
from flask_httpauth import HTTPBasicAuth
import pulp
from werkzeug.wrappers import response
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment

app = Flask(__name__)
auth = HTTPBasicAuth()

# ベーシック認証のためのidとパスワード
# {"ユーザー名": "パスワード"}
id_list = {"test": "0000"}

# headerが1の場合と2の場合を表す変数
header = 1

# ダウンロードするファイルが入る変数
excelFile = None

# 入力されたidに該当するパスワードを比較
@auth.get_password
def get_pw(id):
    if id in id_list:
        return id_list.get(id)
    return None

@app.route('/', methods=['POST', 'GET'])
@auth.login_required # ここで認証が行われる
def index():

    # ここから初期化処理 残ったxlsxファイルとcssファイルを消している
    print("------初期化処理開始------")
    deleteFileXlsx = glob.glob('./*.xlsx')
    deleteFileCss = glob.glob('./*.csv')
    for filename in deleteFileXlsx:
        os.remove(filename)
        print(filename,"を削除しました。")
    for filename in deleteFileCss:
        os.remove(filename)
        print(filename,"を削除しました。")
    print("------初期化処理終了------")
    # ここまで初期化処理

    if request.method == "GET":
        print("index.html: GETでした！")# デバッグ用
        return render_template("index.html",note = 0)

    print("index.html: POSTでした！")# デバッグ用
    # index.htmlの<input type="file" name="csv" class="form-control" id="customFile" accept=".csv">から取得
    f = request.files['csv']

    # もし、ファイル選択に何も指定されなかったら
    if f.filename == "":
        return render_template("index.html",note = 1)

    f.save(f.filename) # ファイルを保存(ファイルを選択して「シフト作成」ボタンを押すとstudio codeの左のファイルマネージャにcsvファイルが表示されるはず)

    try:
        # csvファイルをデータフレームに
        chouseisan_csv = pd.read_csv(f.filename, encoding='cp932', header=1)
        chouseisan_csv['日程'].tolist()
        chouseisan_csv = chouseisan_csv.iloc[: , :-1]
        header = 1
        print("header1で読み込みます")
    except:
        chouseisan_csv = pd.read_csv(f.filename, encoding='cp932', header=2)
        header = 2
        print("header2で読み込みます")

    # ここから曜日を1 0 であらわす処理 ↓

    # タイトルを取得
    with open(f.filename, encoding='cp932') as file:
        title = file.readlines()[0]
        title = title.replace( '\n' , '' )
        title = title.replace( ',' , '')

    # csvファイルの日程の列をリスト化
    day_of_week_list = chouseisan_csv['日程'].tolist()
    
    # 日程のリストを0, 1に変換
    for i,string in enumerate(day_of_week_list):
        # 10/1(金) 11:00～17:00という日程の形から日付を摘出
        target = '('
        idx = string.find(target)
        r = string[:idx]

        try:
            # 摘出した文字列(10/1)をdatetime型に変換
            dte = datetime.datetime.strptime(r, '%m/%d')
            today = datetime.date.today()
            # 調整さんには年は記述されていないので現在の年を追加
            dte = dte.replace(year = today.year)

            
            
            # もし土日、祝日(jpholidayを使用)だったら
            if dte.weekday() >= 5 or jpholiday.is_holiday(dte):
                day_of_week_list[i] = "1"
            else:
                day_of_week_list[i] = "0"

        except:
            """
            日程が
            10/1(金) 11:00～17:00
            17:00～22:00
            このように（2行）なっている場合、二つ目の行にも同じ値を追加
            """
            if day_of_week_list[i - 1] == "1":
                day_of_week_list[i] = "1"
            elif day_of_week_list[i - 1] == "0":
                day_of_week_list[i] = "0"
            else:
                continue

    # 作成したリストをdataflameに追加
    chouseisan_csv.insert(loc = 1, column= '曜日' ,value= day_of_week_list)
    print("曜日を追加したcsvファイル(chouseisan_csv): ",chouseisan_csv) #デバッグ用
    
    # daysとmemberの取得
    days = (len(chouseisan_csv.axes[0]) - 1) // 2 # 提出された表から日数を取得、各日2列なので2で割る
    member = len(chouseisan_csv.axes[1]) - 2 # 提出された表から人数取得

    # シフト希望の○×を0,1に変換
    shift_converted = np.ones((days * 2, member)) # シフトの0,1を格納する箱を作成、全て1が格納されている
    shift_hope = chouseisan_csv.iloc[0:days * 2, 2:] # 調整さんのデータフレームから○×だけを取得
    shift_hope.to_string(header=False, index=False) # ヘッダーとインデックスの削除、○×だけの状態に

    for i in (range(days * 2)): # 日程の分ループさせる
        for j in (range(member)): # 従業員の分ループさせる
            if shift_hope.iat[i, j] != "○": # もしシフト希望表のあるマスが×ならそのマスに0を格納
                shift_converted[i, j] = 0
    



    print("○×を1,0に置き換えたもの(shift_converted): ",shift_converted) # ○×が1,0に書き換えられたシフト希望表の2次元配列を出力

    needNumberWeekday = [3, 2] # [前半, 後半]
    needNumberHoliday = [4, 4] # [前半, 後半]

    # ペナルティ定数の定義
    C_needNumber = 10

    # 問題の定義
    problem = pulp.LpProblem(name="penalty", sense=pulp.LpMinimize)

    # 変数の定義
    V_shift = np.array(addbinvars(days*2, member))
    
    V_needNumber = np.array(addbinvars(days))

    # 必要な条件
    # ・×が提出されている人をアサインしてはいけない
    # ・○のみでシフトを作る
    # ・平日、前半は2人後半は1人
    # 　休日、前半は3人後半3人
    # ・給料の誤差少ないようにする

    # 目的関数
    problem += C_needNumber * pulp.lpSum(V_needNumber)
    #    + C_noAssign * lpSum(V_noAssign)

    weekday = "0"
    holiday = "1"
    
    no_shift_hope = 0.0
    yes_shift_hope = 1.0

    # 制約関数
    # シフト希望が×ならシフトに入れない
    for i in range(days * 2):
        for j in range(member):
            if shift_converted[i][j] == no_shift_hope:
                problem += V_shift[i][j] == 0
            elif shift_converted[i][j] == yes_shift_hope:
                continue
            else:
                print("155行目:×の人にはアサインしないようにするループでエラー発生")
                break

    # 人数が合うようにする制約式
    for i in range(0, days * 2, 2):
        # if chouseisan_csvが×ならそこに0を入れる制約式を作る
        if chouseisan_csv.iloc[i,1] == weekday:
            problem += V_needNumber[i//2] >= (pulp.lpSum(V_shift[i][j] for j in range(member)) - needNumberWeekday[0])
            problem += V_needNumber[i//2] >= -(pulp.lpSum(V_shift[i][j] for j in range(member)) - needNumberWeekday[0])
            problem += V_needNumber[i//2] >= (pulp.lpSum(V_shift[i+1][j] for j in range(member)) - needNumberWeekday[1])
            problem += V_needNumber[i//2] >= -(pulp.lpSum(V_shift[i+1][j] for j in range(member)) - needNumberWeekday[1])
        elif chouseisan_csv.iloc[i,1] == holiday:
            problem += V_needNumber[i//2] >= (pulp.lpSum(V_shift[i][j] for j in range(member)) - needNumberHoliday[0])
            problem += V_needNumber[i//2] >= -(pulp.lpSum(V_shift[i][j] for j in range(member)) - needNumberHoliday[0])
            problem += V_needNumber[i//2] >= (pulp.lpSum(V_shift[i+1][j] for j in range(member)) - needNumberHoliday[1])
            problem += V_needNumber[i//2] >= -(pulp.lpSum(V_shift[i+1][j] for j in range(member)) - needNumberHoliday[1])
        else:
            print("実行不可", chouseisan_csv.iloc[i,1])

    # 解く
    print("------計算します------")
    status = problem.solve()
    print("pulpステータス: ",pulp.LpStatus[status])


    result = np.vectorize(pulp.value)(V_shift).astype(int)
    print("作成されたシフト(result): ",result) # 作成されたシフト


    # デバッグ用
    for i in range(days * 2):
        for j in range(member):
            if shift_converted[i][j] == no_shift_hope:
                if result[i, j] != 0:
                    print((i,j) , "シフト希望不可の部分にアサインしてしまっているので０に変えます。")
                    result[i, j] = 0
                    print(result[i, j])
            else:
                continue
    
    print("直されたシフト(result): ",result) # 直されたシフト

    # 結果表示
    print("制約関数",pulp.value(problem.objective))

    # 作成されたシフトをエクセルで出力する
    # もう一度csvを読み込んでその中の○×を書き換える
    if header == 1:
        new_chouseisan_csv = pd.read_csv(f.filename, encoding='cp932', header=1)
        new_chouseisan_csv = new_chouseisan_csv.iloc[: , :-1]
    else:
        new_chouseisan_csv = pd.read_csv(f.filename, encoding='cp932', header=2)

     # csv読み込み
    new_chouseisan_csv = new_chouseisan_csv.fillna("") # 欠損値(Nan)を消す
    # title = new_chouseisan_csv.iat[0, 0] # ファイルの名前にする部分を取得 
    # ↑ここを取得するのはまだ

    # エクセルファイルに変換する処理
    for i in (range(days * 2)): # 日程の分ループさせる
        for j in (range(member)): # 従業員の分ループさせる
            if result[i, j] == 1: # もしシフトのあるマスが1ならそのマスに○を格納

                new_chouseisan_csv.iat[i, j + 1] = "○"
            elif result[i, j] == 0: # もしシフトのあるマスが0ならそのマスに×を格納
                new_chouseisan_csv.iat[i, j + 1] = "×"
            else: # 0 or 1 以外がある場合,シフト希望を見て○×を選択
                if shift_converted[i, j] == no_shift_hope:
                    new_chouseisan_csv.iat[i, j + 1] = "×"
                elif shift_converted[i, j] == yes_shift_hope:
                    new_chouseisan_csv.iat[i, j + 1] = "○"
                else:
                    new_chouseisan_csv.iat[i, j + 1] = "error"

    print("エクセルファイルの中身: ",new_chouseisan_csv) # エクセルファイルの中身
    new_chouseisan_csv.to_excel(title + '.xlsx', encoding='cp932', index=False, header=True) #インデックス、ヘッダーなしでエクセル出力

    # この二つの変数がグローバル変数であることの定義
    global excelFile

    excelFile = title + '.xlsx'
    os.remove(f.filename) # 処理が終わった後、ダウンロードしたcsvを消す   

    # ここからエクセルファイル編集 希望しているところに色付け
    
    wb = openpyxl.load_workbook(filename=excelFile)
    sheet = wb['Sheet1']

    print(type(shift_hope))
    print(shift_hope)

    # B列に人数列を挿入する
    sheet.insert_cols(2)
    sheet["B1"].value = "人数"
    sheet["B1"].alignment = Alignment( # 中央揃えに変更
        horizontal='center',
        vertical='center',
    )

    # B列に=countif関数を挿入
    for i in range(2, days*2+2):
        countif_people_number = "=COUNTIF(C" + str(i) + ":" + chr(member+66) + str(i) + ',"○")'
        sheet.cell(row=i, column=2).value = countif_people_number

    for i in(range(days * 2)):
        for j in (range(member)):
            if shift_hope.iat[i,j] == "○":
                # シフト希望を出しているところに色付け
                sheet.cell(row=i+2, column=j+3).fill = PatternFill(patternType='solid', fgColor='fecf8d', bgColor= 'fecf8d') # オレンジ
            elif shift_hope.iat[i,j] == "×":
                continue
            else:
                sheet.cell(row=i+2, column=j+3).fill = PatternFill(patternType='solid', fgColor='fffac2', bgColor= 'fffac2') # 黄色
    
    # 変更したエクセルファイルを変更
    wb.save(excelFile)

    #　希望しているところに色付けここまで
    
    # 表のmemberと日程を見えやすいように色付け
    for i in range(1, days*2+4):
        sheet.cell(row=i, column=1).fill = PatternFill(patternType='solid', fgColor='eaf6fd', bgColor= 'eaf6fd') # 水色

    for i in range(1, member+3):
        sheet.cell(row=1, column=i).fill = PatternFill(patternType='solid', fgColor='eaf6fd', bgColor= 'eaf6fd') # 水色

    # 人が不足している場合の色付け

    fill_red = PatternFill(start_color='f69679', end_color='f69679', fill_type='solid')
    dxf=DifferentialStyle(fill=fill_red)

    for i in range(2, days * 2 + 2):
        rule_Weekday = Rule(type='expression', formula=['B' + str(i) + '<2'], dxf=dxf)
        rule_Holiday = Rule(type='expression', formula=['B' + str(i) + '<4'], dxf=dxf)
        if chouseisan_csv.iloc[i-2,1] == weekday:
            sheet.conditional_formatting.add('B' + str(i), rule_Weekday)
        elif chouseisan_csv.iloc[i-2,1] == holiday:
            sheet.conditional_formatting.add('B' + str(i), rule_Holiday)
    
    # 人が不足している場合の色付けここまで

    # 変更したエクセルファイルを変更
    wb.save(excelFile)
    
    # A列の幅を広くする
    sheet.column_dimensions['A'].width = 23

    # ズーム倍率を140%に変更
    sheet.sheet_view.zoomScale = 140

    # countifで○の数を数える
    sheet["A" + str(days*2+3)].value = "予想給料"
    
    for i in range(1,member+1):
        countif_circle = "=COUNTIF(" + chr(i+66) + "2:" + chr(i+66) + str(days*2+1) + ',"○")*5000 &"円"'
        sheet.cell(row=days*2+3, column=i+2).value = countif_circle

    # ○,×,△のプルダウンを作成
    dv = DataValidation(type="list", formula1='"○,×,△"')

    # 適用するセルの指定
    for i in range(1, days*2+1):
        for j in range(1, member+1):
            dv.add(sheet.cell(i+1, j+2))
    
    # シートに入力規則を登録
    sheet.add_data_validation(dv)

    # 罫線(外枠)を設定
    thin_bottom_border = Border(bottom=Side(style='thin', color='000000'))
    medium_bottom_border = Border(bottom=Side(style='medium', color='000000'))

    # セルに罫線を設定
    # 1行目の下側に中線
    for i in range(1, member+2):
        sheet.cell(row=1, column=i+1).border = medium_bottom_border
    
    # 日程ごとに下側に細線
    for i in range(3, days*2+3, 2):
        for j in range(1, member+2):
            sheet.cell(row=i, column=j+1).border = thin_bottom_border
    
    # 1行目を画面に固定
    sheet.freeze_panes = 'A2'

    # 変更したエクセルファイルを変更
    wb.save(excelFile)
    
    # 結果用のhtml
    return render_template("finished.html", file = title + '.xlsx')

@app.route('/download', methods=['POST', 'GET'])
def download():

    print("ダウンロードされたファイル: ",excelFile)

    return send_file(excelFile, as_attachment=True,
                     attachment_filename=excelFile,
                     mimetype='text/plain')


if __name__ == '__main__':
    app.debug = True
    
    app.run()